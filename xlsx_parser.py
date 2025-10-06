from openpyxl import load_workbook
from openpyxl.cell import MergedCell
from datetime import datetime, timedelta
import re

class ScheduleParser:
    def __init__(self, xlsx_path):
        self.workbook = load_workbook(xlsx_path)
        self.sheet = self.workbook.active

    def parse(self):
        events = []
        time_slots = self._parse_time_slots()

        # Start from row 4 (after header rows)
        for row_idx in range(4, self.sheet.max_row + 1):
            date_cell = self.sheet[f'A{row_idx}']

            # Skip empty rows
            if not date_cell.value:
                continue

            date_info = self._parse_date_info(date_cell.value)
            if not date_info:
                continue

            # Parse events in this row
            for col_idx, time_slot in enumerate(time_slots, start=2):  # B column onwards
                cell = self.sheet.cell(row=row_idx, column=col_idx)

                if isinstance(cell, MergedCell):
                    continue

                if cell.value:
                    event = self._parse_event(cell.value, date_info, time_slot, row_idx, col_idx)
                    if event:
                        events.append(event)

        return events

    def _parse_time_slots(self):
        # Extract time slots from header row (row 3)
        time_slots = []
        for col_idx in range(2, self.sheet.max_column + 1):  # Start from column B
            cell = self.sheet.cell(row=3, column=col_idx)
            if cell.value:
                time_slots.append(str(cell.value).strip())
        return time_slots

    def _parse_date_info(self, date_str):
        # Parse date string
        if not date_str:
            return None

        date_str = str(date_str).strip()
        is_online = 'ONLINE' in date_str.upper()

        # Extract date pattern
        date_match = re.search(r'(\d{1,2})\.(\d{1,2})\.(\d{4})', date_str)
        if not date_match:
            return None

        day, month, year = date_match.groups()
        date = datetime(int(year), int(month), int(day))

        return {
            'date': date,
            'is_online': is_online
        }

    def _parse_event(self, event_str, date_info, time_slot, row_idx, col_idx):

        event_str = str(event_str).strip()

        # Check how many columns this cell spans/merged
        col_span = self._get_merged_cols(row_idx, col_idx)

        # Parse event details
        # Format: "Subject - type - Professor Classroom"
        parts = event_str.split(' - ')
        if len(parts) < 2:
            return None

        subject = parts[0].strip()
        event_type = parts[1].strip() if len(parts) > 1 else ''

        # Extract professor
        professor = ''
        classroom = ''
        if len(parts) > 2:
            last_part = parts[2].strip()
            # Classroom is usually at the end (e.g., B401)
            classroom_match = re.search(r'([A-Z]\d+)\s*$', last_part)
            # Extract classroom if found
            if classroom_match:
                classroom = classroom_match.group(1)
                professor = last_part[:classroom_match.start()].strip()
            else:
                professor = last_part

        start_time, end_time = self._parse_time_slot(time_slot, col_span)
        if not start_time:
            return None

        # Combine date and time
        start_datetime = datetime.combine(date_info['date'], start_time)
        end_datetime = datetime.combine(date_info['date'], end_time)

        return {
            'subject': subject,
            'type': event_type,
            'professor': professor,
            'classroom': classroom,
            'start': start_datetime,
            'end': end_datetime,
            'is_online': date_info['is_online']
        }

    """helper functions"""

    def _get_merged_cols(self, row_idx, col_idx):
        """Get how many columns a merged cell spans"""
        for merged_range in self.sheet.merged_cells.ranges:
            if (merged_range.min_row <= row_idx <= merged_range.max_row and
                    merged_range.min_col <= col_idx <= merged_range.max_col):
                return merged_range.max_col - merged_range.min_col + 1
        return 1

    def _parse_time_slot(self, time_slot, col_span=1):
        """Parse time slot like '10.30-11.15' and extend based on column span"""
        match = re.search(r'(\d{1,2})\.(\d{2})-(\d{1,2})\.(\d{2})', time_slot)
        if not match:
            return None, None

        start_hour, start_min, end_hour, end_min = map(int, match.groups())
        start_time = datetime.strptime(f"{start_hour}:{start_min}", "%H:%M").time()

        # If cell spans multiple columns, extend the end time
        base_end = datetime.strptime(f"{end_hour}:{end_min}", "%H:%M")
        # Each column is typically 45 minutes
        extended_end = base_end + timedelta(minutes=45 * (col_span - 1))
        end_time = extended_end.time()

        return start_time, end_time

    """main function"""

if __name__ == "__main__":
    # Test the parser
    parser = ScheduleParser("schedule.xlsx")
    events = parser.parse()

    for event in events:
        print(f"{event['start']} - {event['subject']} ({event['type']}) - {event['professor']} - {event['classroom']}")